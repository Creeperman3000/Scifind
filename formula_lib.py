"""
Shared library for the physics formula cheat-sheet.
Provides DB connection, LaTeX rendering, dimension formatting,
SI unit parsing, and CSV import/export.
"""

import csv
import html
import json
import math
import os
import sqlite3
from collections import defaultdict
from fractions import Fraction
from io import StringIO
from pathlib import Path

DEFAULT_DB = str(Path.home() / ".local" / "share" / "formula" / "formulas.db")

DIM_ORDER = ["M", "L", "T", "I", "Θ", "N", "J"]
DIM_COLS = ["dim_M", "dim_L", "dim_T", "dim_I", "dim_Θ", "dim_N", "dim_J"]
DIM_VAR_IDS = {"M": "mass", "L": "length", "T": "period", "I": "electric_current",
               "Θ": "temperature", "N": "amount", "J": "luminous_intensity"}


def db_path():
    return os.environ.get("FORMULA_DB", DEFAULT_DB)


def get_conn():
    p = db_path()
    os.makedirs(os.path.dirname(p), exist_ok=True)
    conn = sqlite3.connect(p)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    return conn


def en(data, locale="en-us"):
    if not data:
        return ""
    try:
        d = json.loads(data)
        for key in (locale, "en-us"):
            val = d.get(key)
            if val:
                return val
        return str(data)
    except (json.JSONDecodeError, TypeError):
        return str(data)


def rebuild_fts(conn):
    conn.execute("DELETE FROM formula_fts")
    rows = conn.execute("""
        SELECT f.id, f.name, f.description,
               COALESCE(group_concat(v.name_en, ' '), '') AS vars
        FROM formulas f
        LEFT JOIN formula_items fi ON fi.formula_id = f.id
        LEFT JOIN (SELECT DISTINCT id, json_extract(name, '$.en-us') AS name_en
                   FROM variables) v ON v.id = fi.variable_id
        GROUP BY f.id
    """).fetchall()
    for r in rows:
        conn.execute(
            "INSERT INTO formula_fts (formula_id, name, description, variables) VALUES (?, ?, ?, ?)",
            (r["id"], r["name"], r["description"], r["vars"]),
        )

    conn.execute("DELETE FROM variable_fts")
    vrows = conn.execute("""
        SELECT id, json_extract(name, '$.en-us') AS name_en, latex
        FROM variables
    """).fetchall()
    for r in vrows:
        conn.execute(
            "INSERT INTO variable_fts (variable_id, name, latex) VALUES (?, ?, ?)",
            (r["id"], r["name_en"], r["latex"]),
        )

    conn.execute("DELETE FROM unit_fts")
    urows = conn.execute("""
        SELECT id, json_extract(name, '$.en-us') AS name_en, symbol
        FROM units
    """).fetchall()
    for r in urows:
        conn.execute(
            "INSERT INTO unit_fts (unit_id, name, symbol) VALUES (?, ?, ?)",
            (r["id"], r["name_en"], r["symbol"]),
        )

    conn.commit()
    return len(rows)


def migrate_db(conn):
    """Apply schema migrations for existing databases."""
    cols = [r[1] for r in conn.execute("PRAGMA table_info(units)").fetchall()]
    if "name" not in cols:
        conn.execute("ALTER TABLE units ADD COLUMN name TEXT NOT NULL DEFAULT '{}'")
        for row in conn.execute("SELECT id FROM units").fetchall():
            uid = row["id"]
            en_name = uid.replace("_", " ").title()
            conn.execute("UPDATE units SET name = ? WHERE id = ?",
                         (json.dumps({"en": en_name}), uid))
        conn.commit()

    # Migrate old JSON i18n {"en":"..."} to {"en-us":"..."}
    for table in ("formulas", "variables", "units", "conditions"):
        # Check if the table has a name column
        tcols = [r[1] for r in conn.execute(f"PRAGMA table_info({table})").fetchall()]
        json_cols = []
        for c in ("name", "science", "branch", "topic", "description"):
            if c in tcols:
                json_cols.append(c)
        for col in json_cols:
            conn.execute(f"""
                UPDATE {table}
                SET {col} = json_set({col}, '$.en-us', json_extract({col}, '$.en'))
                WHERE json_extract({col}, '$.en') IS NOT NULL
                  AND json_extract({col}, '$.en-us') IS NULL
            """)
            # Remove the old $.en key
            conn.execute(f"""
                UPDATE {table}
                SET {col} = json_remove({col}, '$.en')
                WHERE json_extract({col}, '$.en-us') IS NOT NULL
                  AND json_extract({col}, '$.en') IS NOT NULL
            """)
    conn.commit()

    # Add en-uk locale keys for known US/UK spelling differences in units
    UK_OVERRIDES = {
        "Meter": "Metre", "Centimeter": "Centimetre", "Decimeter": "Decimetre",
        "Liter": "Litre", "Center of Mass": "Centre of Mass",
    }
    for row in conn.execute("SELECT id, name FROM units WHERE json_extract(name, '$.en-uk') IS NULL").fetchall():
        try:
            d = json.loads(row["name"])
            en_us = d.get("en-us", "")
            for us_spelling, uk_spelling in UK_OVERRIDES.items():
                if us_spelling in en_us:
                    uk_val = en_us.replace(us_spelling, uk_spelling)
                    if uk_val != en_us:
                        conn.execute(
                            "UPDATE units SET name = json_set(name, '$.en-uk', ?) WHERE id = ?",
                            (uk_val, row["id"]),
                        )
                        break
        except (json.JSONDecodeError, TypeError):
            pass
    conn.commit()

    # Migrate old-style LaTeX symbols to siunitx-compatible ones
    symbol_migrations = {
        "\\\\text{\\\\textdegree C}": "\\degreeCelsius",
        "\\\\Omega": "\\ohm",
        "N": "\\newton",
    }
    for old, new in symbol_migrations.items():
        conn.execute("UPDATE units SET symbol = ? WHERE symbol = ?", (new, old))
    conn.commit()

    # Ensure FTS tables exist (for databases created before search was expanded)
    for fts_sql in [
        "CREATE VIRTUAL TABLE IF NOT EXISTS variable_fts USING fts5(variable_id UNINDEXED, name, latex)",
        "CREATE VIRTUAL TABLE IF NOT EXISTS unit_fts USING fts5(unit_id UNINDEXED, name, symbol)",
    ]:
        try:
            conn.execute(fts_sql)
        except Exception:
            pass
    conn.commit()

    # Seed new variables and units (derived quantities, non-SI units)
    _SCRIPT_DIR = Path(__file__).parent
    units_seed = _SCRIPT_DIR / "seed_units.sql"
    if units_seed.exists():
        conn.executescript(units_seed.read_text())
        conn.commit()


def fmt_num(n):
    if n == int(n):
        return str(int(n))
    s = f"{n:.10f}".rstrip("0")
    return s.rstrip(".")


# ── Dimension rendering ────────────────────────────────

def render_dimensions(dim_M=0, dim_L=0, dim_T=0, dim_I=0, dim_Theta=0, dim_N=0, dim_J=0):
    """Render dimension exponents as human-readable string."""
    vals = [dim_M, dim_L, dim_T, dim_I, dim_Theta, dim_N, dim_J]
    parts = []
    for k, exp in zip(DIM_ORDER, vals):
        if exp is None:
            exp = 0
        if exp != 0:
            p = k
            if exp != 1:
                p += f"^{fmt_num(exp)}"
            parts.append(p)
    return " \u00b7 ".join(parts) if parts else "dimensionless"


def render_dimensions_latex(dim_M=0, dim_L=0, dim_T=0, dim_I=0, dim_Theta=0, dim_N=0, dim_J=0,
                            var_latex_map=None, unit_symbol_map=None, mode="variables"):
    """Render dimension exponents as LaTeX.

    var_latex_map: dict mapping variable_id → LaTeX string from variables.latex
    unit_symbol_map: dict mapping variable_id → unit symbol for LaTeX from units.symbol
    mode: "variables" (default, uses var_latex_map) or "units" (uses unit_symbol_map)
    """
    vals = [dim_M, dim_L, dim_T, dim_I, dim_Theta, dim_N, dim_J]
    lookup = var_latex_map if (mode == "var" and var_latex_map) else unit_symbol_map
    parts = []
    for dim_sym, exp in zip(DIM_ORDER, vals):
        if not exp:
            continue
        var_id = DIM_VAR_IDS[dim_sym]
        sym = lookup.get(var_id) if lookup else None
        if sym is None:
            sym = f"\\mathrm{{{dim_sym}}}"
        if exp == 1:
            parts.append(sym)
        else:
            e = str(int(exp)) if exp == int(exp) else str(exp)
            parts.append(f"{sym}^{{{e}}}")
    if not parts:
        return "\\text{dimensionless}"
    return " \\cdot ".join(parts)


def dims_from_row(row):
    """Extract dimension values from a row (dict or sqlite3.Row)."""
    cols = ["dim_M", "dim_L", "dim_T", "dim_I", "dim_Θ", "dim_N", "dim_J"]
    return [dict(row).get(c, 0) or 0 for c in cols]


# ── SI unit rendering ──────────────────────────────────

def parse_si_unit_json(si_unit_json):
    """Parse si_unit JSON and return list of (unit_id, exponent)."""
    if not si_unit_json:
        return []
    try:
        parts = json.loads(si_unit_json)
        return [(p["unit"], p["exponent"]) for p in parts]
    except (json.JSONDecodeError, KeyError, TypeError):
        return []


def decompose_si_unit_parts(si_unit_json):
    """Parse si_unit JSON and decompose composite units (e.g. metre_per_second) into base parts.

    Returns list of (unit_id, exponent) with composites expanded.
    """
    parts = parse_si_unit_json(si_unit_json)
    result = []
    for uid, exp in parts:
        comp = unit_id_components(uid)
        if comp is None:
            result.append((uid, exp))
        else:
            nums, dens = comp
            for nid, nexp in nums:
                result.append((nid, nexp * exp))
            if dens:
                for did, dexp in dens:
                    result.append((did, -dexp * exp))
    return result


def render_si_unit_html(si_unit_json, unit_url_func=None, unit_name_func=None, locale="en-us"):
    """Render si_unit JSON as HTML with links."""
    parts = decompose_si_unit_parts(si_unit_json)
    if not parts:
        return ""
    ls = _get_ls(locale)
    num_parts = []
    den_parts = []
    for uid, exp in parts:
        if exp >= 0:
            num_parts.append((uid, exp))
        else:
            den_parts.append((uid, -exp))
    num_str = _render_unit_group(num_parts, unit_url_func, unit_name_func, locale)
    if not den_parts:
        return num_str
    den_str = _render_unit_group(den_parts, unit_url_func, unit_name_func, locale)
    if not num_str:
        return f"{ls['reciprocal']} {den_str}"
    return f"{num_str} {ls['per']} {den_str}"


def _render_unit_group(parts, url_func, name_func=None, locale="en-us"):
    """Render a list of (unit_id, exponent) as HTML with natural-language exponents."""
    items = []
    for uid, exp in parts:
        label = name_func(uid) if name_func else uid.replace("_", " ").title()
        word = _exp_word(exp, locale)
        if url_func:
            link = url_func(uid)
            text = f'<a href="{html.escape(link)}">{html.escape(label)}</a>'
        else:
            text = html.escape(label)
        if word:
            text += " " + html.escape(word)
        items.append(text)
    return "-".join(items)


# ── Unit ID decomposition ──────────────────────────────

def is_composite_unit(unit_id):
    """Return True if the unit ID describes a derived/composite unit."""
    return (
        "_per_" in unit_id
        or unit_id.startswith("square_")
        or unit_id.startswith("cubic_")
        or unit_id.startswith("reciprocal_")
    )

def _strip_unit_suffix(uid):
    """Remove disambiguation suffixes from unit IDs (e.g. _mass, _energy)."""
    for suffix in ("_mass", "_energy", "_luminous", "_mechanical", "_thermal",
                   "_gravitational", "_molar", "_specific", "_radiant"):
        if uid.endswith(suffix):
            return uid[:-len(suffix)]
    return uid


def _parse_unit_part(part_str):
    """Parse a numerator/denominator part string into [(unit_id, exponent)]."""
    # Suffixes contain underscores that split() would eat, so replace
    # them with markers before splitting so "metre_sq" → ["metre","\x00sq"].
    SUFFIXES = ["_sq"] + [f"_{s}" for s in
        ("mass", "energy", "luminous", "mechanical", "thermal",
         "gravitational", "molar", "specific", "radiant")]
    markers = {}
    for i, sfx in enumerate(SUFFIXES):
        marker = f"\x00{i}"
        markers[marker] = sfx
        part_str = part_str.replace(sfx, marker)

    tokens = part_str.split("_")
    result = []
    i = 0
    while i < len(tokens):
        tok = tokens[i]
        if not tok:
            i += 1
            continue
        exp = 1
        if tok == "square":
            exp = 2
            i += 1
            if i < len(tokens):
                tok = tokens[i]
            else:
                break
        elif tok == "cubic":
            exp = 3
            i += 1
            if i < len(tokens):
                tok = tokens[i]
            else:
                break
        elif tok == "reciprocal":
            exp = -1
            i += 1
            if i < len(tokens):
                tok = tokens[i]
            else:
                break
        # Strip any suffix markers (_sq → exp=2, _mass → strip, etc.)
        for marker, orig in sorted(markers.items(), key=lambda x: -len(x[0])):
            if tok == marker:
                tok = ""
                break
            if tok.endswith(marker):
                tok = tok[: -len(marker)]
                if orig == "_sq":
                    exp = 2
                break
        if not tok:
            i += 1
            continue
        tok = _strip_unit_suffix(tok)
        result.append((tok, exp))
        i += 1
    return result


def unit_id_components(unit_id):
    """Parse a composite unit ID into (num_parts, den_parts).

    Returns (num_units, den_units) or None for a simple (non-composite) unit.
    Each part is a list of (base_unit_id, exponent).
    """
    parts = unit_id.split("_per_")
    if len(parts) == 1:
        pid = parts[0]
        if any(pid.startswith(p) for p in ("square_", "cubic_", "reciprocal_")):
            nums = _parse_unit_part(pid)
            return (nums, None)
        return None  # Simple unit

    num_str = parts[0]
    den_str = "_per_".join(parts[1:])  # In case there are multiple _per_ (unlikely)
    nums = _parse_unit_part(num_str)
    dens = _parse_unit_part(den_str)
    return (nums, dens)


LOCALE_STRINGS = {
    "en-us": {
        "squared": "squared",
        "cubed": "cubed",
        "inverse": "inverse",
        "to_the": "to the",
        "per": "per",
        "reciprocal": "Reciprocal",
        "dimensionless": "dimensionless",
        "exp_prefix": "",
    },
    "en-uk": {
        "squared": "squared",
        "cubed": "cubed",
        "inverse": "inverse",
        "to_the": "to the",
        "per": "per",
        "reciprocal": "Reciprocal",
        "dimensionless": "dimensionless",
        "exp_prefix": "",
    },
}

ORDINAL_SPECIAL = {11: "th", 12: "th", 13: "th"}
ORDINAL_LAST_DIGIT = {0: "th", 1: "st", 2: "nd", 3: "rd",
                      4: "th", 5: "th", 6: "th", 7: "th", 8: "th", 9: "th"}


def _get_ls(locale):
    """Get locale strings, falling back to en-us."""
    return LOCALE_STRINGS.get(locale, LOCALE_STRINGS["en-us"])


def _ordinal(n):
    last2 = n % 100
    if last2 in ORDINAL_SPECIAL:
        suffix = ORDINAL_SPECIAL[last2]
    else:
        suffix = ORDINAL_LAST_DIGIT[n % 10]
    return f"{n}{suffix}"


def _exp_word(exp, locale="en-us"):
    """Return natural-language exponent word for a unit."""
    ls = _get_ls(locale)
    if exp == 2:
        return ls["squared"]
    elif exp == 3:
        return ls["cubed"]
    elif exp == 1:
        return ""
    elif exp == -1:
        return ls["inverse"]
    elif exp > 3:
        return f"{ls['to_the']} {_ordinal(exp)}"
    else:
        return ""


def render_unit_decomposition(unit_id, name_func, url_func=None, locale="en-us"):
    """Render a unit name with links, decomposing composite units into components.

    Simple units (metre, newton) are returned as linked names.
    Composite units (metre_per_second) are decomposed into linked components.
    Exponents use natural language (squared, cubed, to the Nth).
    """
    comp = unit_id_components(unit_id)
    if comp is None:
        name = name_func(unit_id) if name_func else unit_id.replace("_", " ").title()
        if url_func:
            url = url_func(unit_id)
            return f'<a href="{html.escape(url)}">{html.escape(name)}</a>'
        return html.escape(name)

    nums, dens = comp

    # Move negative-exponent numerator parts to denominator
    fixed_nums = []
    fixed_dens = list(dens) if dens else []
    for uid, exp in nums:
        if exp < 0:
            fixed_dens.append((uid, -exp))
        else:
            fixed_nums.append((uid, exp))

    ls = _get_ls(locale)
    if fixed_nums and fixed_dens:
        num_html = _render_unit_group(fixed_nums, url_func, name_func, locale)
        den_html = _render_unit_group(fixed_dens, url_func, name_func, locale)
        return f"{num_html} {ls['per']} {den_html}"
    elif fixed_dens:
        den_html = _render_unit_group(fixed_dens, url_func, name_func, locale)
        return f"{ls['reciprocal']} {den_html}"
    return _render_unit_group(fixed_nums, url_func, name_func, locale)


# ── LaTeX rendering helpers ────────────────────────────

def _is_integer(n):
    if n is None:
        return True
    return n == int(n)


def _render_exp(exp, base_str):
    """Render a variable/coeff part with exponent handling.

    For integer exponents: var^{exp} or just var for exp=1.
    For negative integer exponents: frac{1}{var^{abs(exp)}}
    For fractional exponents: sqrt[den]{var^{num}}
    """
    if exp is None or exp == 1:
        return base_str

    if _is_integer(exp):
        iexp = int(exp)
        if iexp < 0:
            inner = base_str
            if iexp != -1:
                inner += "^{" + fmt_num(-iexp) + "}"
            return "\\frac{1}{" + inner + "}"
        return base_str + "^{" + fmt_num(iexp) + "}"

    # fractional exponent → root
    try:
        f = Fraction(exp).limit_denominator(100)
        num = f.numerator
        den = f.denominator
    except Exception:
        return base_str + "^{" + fmt_num(exp) + "}"

    if num == 1:
        if den == 2:
            return "\\sqrt{" + base_str + "}"
        return "\\sqrt[" + str(den) + "]{" + base_str + "}"
    if num == -1:
        if den == 2:
            return "\\frac{1}{\\sqrt{" + base_str + "}}"
        return "\\frac{1}{\\sqrt[" + str(den) + "]{" + base_str + "}}"
    inner = base_str + "^{" + fmt_num(num) + "}"
    if den == 2:
        return "\\sqrt{" + inner + "}"
    return "\\sqrt[" + str(den) + "]{" + inner + "}"


def render_variable(item, flipped):
    var = item["latex_override"] or item["var_latex"] or item["variable_id"] or "?"
    prefix = item["latex_prefix"] or ""
    suffix = item["latex_suffix"] or ""

    if prefix:
        var = prefix + var
    if suffix:
        var = var + suffix

    if item["label"]:
        var += "_{" + item["label"] + "}"

    exp = item["var_exponent"] if item["var_exponent"] is not None else 1
    if flipped:
        exp = -exp

    return _render_exp(exp, var)


def render_coeff(item, first_in_term):
    body = None
    is_neg = False
    coeff_special = item["coeff_special"]
    coeff_value = item["coeff_value"]
    coeff_exp = item["coeff_exponent"] if item["coeff_exponent"] is not None else 1

    if coeff_special == "pi":
        body = "\\pi"
    elif coeff_special == "e":
        body = "e"
    elif coeff_value is not None:
        v = coeff_value
        if v < 0:
            is_neg = True
            v = abs(v)
        body = fmt_num(v) if v != 1 else None
    else:
        body = None

    if body is not None:
        body = _render_exp(coeff_exp, body)
        # suppress "1" when it would just be "1"
        if coeff_special is None and body == "1":
            body = None
    elif coeff_exp is not None and coeff_exp != 1:
        body = _render_exp(coeff_exp, "1")

    if first_in_term and is_neg and body in (None, "", "1"):
        body = ""
        is_neg = False

    return body if body else None, is_neg


def _render_items_group(items, flipped):
    """Render a list of items, returning (numerator_str, denominator_str).

    Splits items by sign of the rendered exponent:
    - items with non-negative exponents → numerator
    - items with negative exponents → denominator
    """
    num_parts = []
    den_parts = []
    for i, item in enumerate(items):
        c, neg = render_coeff(item, first_in_term=(i == 0))

        if item["variable_id"]:
            exp = item["var_exponent"] if item["var_exponent"] is not None else 1
            if flipped:
                exp = -exp
            if exp < 0:
                # denominator: make exponent positive
                den_item = dict(item)
                if flipped:
                    den_item["var_exponent"] = -item["var_exponent"] if item["var_exponent"] is not None else -1
                else:
                    den_item["var_exponent"] = item["var_exponent"]
                # We want positive exp in denominator, so flip back
                # For den, exp < 0 means it goes to denominator.
                # render_variable with negative exp will produce \frac{1}{...}
                # We want just the positive part, so set exp positive
                if flipped:
                    den_item["var_exponent"] = -(item["var_exponent"] if item["var_exponent"] is not None else 1)
                else:
                    den_item["var_exponent"] = item["var_exponent"] if item["var_exponent"] is not None else 1
                # Actually, we need to set the exponent to its absolute value
                v = render_variable(dict(item, var_exponent=abs(exp) if not flipped else abs(item["var_exponent"] if item["var_exponent"] is not None else 1)), flipped=False)
                if c:
                    den_parts.append(c)
                den_parts.append(v)
            else:
                v = render_variable(item, flipped=flipped)
                if c:
                    num_parts.append(c)
                num_parts.append(v)
        elif c:
            # coeff-only item — sign already in c
            if neg:
                den_parts.append(c)
            else:
                num_parts.append(c)

    return "".join(num_parts), "".join(den_parts)


def render_formula_items(items):
    """Build LaTeX string from formula_items rows. Handles fractions for negatives."""
    by_term = defaultdict(list)
    for item in items:
        by_term[item["term"]].append(item)

    lhs_terms = []
    rhs_terms = []

    for term_num in sorted(by_term):
        term_items = sorted(by_term[term_num], key=lambda x: (not x["is_primary"], x["sort_order"]))
        primary = [i for i in term_items if i["is_primary"]]
        non_primary = [i for i in term_items if not i["is_primary"]]

        # LHS: primary items (exponents flipped)
        lhs_num, lhs_den = _render_items_group(primary, flipped=True)
        if lhs_num or lhs_den:
            if lhs_den:
                # Need to handle sign for the numerator
                # Check if numerator starts with a minus sign
                has_neg = lhs_num.startswith("-")
                clean_num = lhs_num.lstrip("-")
                if has_neg:
                    lhs_str = f"-\\frac{{{clean_num}}}{{{lhs_den}}}"
                else:
                    lhs_str = f"\\frac{{{lhs_num}}}{{{lhs_den}}}" if lhs_num else f"\\frac{{1}}{{{lhs_den}}}"

                # If the numerator is just a negative coeff like "-", handle it
                if lhs_str == "\\frac{}{}":
                    lhs_str = "\\frac{1}{" + lhs_den + "}"
            else:
                lhs_str = lhs_num
            if lhs_str:
                lhs_terms.append(lhs_str)

        # RHS: non-primary items (exponents as-is)
        rhs_num, rhs_den = _render_items_group(non_primary, flipped=False)
        if rhs_num or rhs_den:
            if rhs_den:
                has_neg = rhs_num.startswith("-")
                clean_num = rhs_num.lstrip("-")
                if has_neg:
                    rhs_str = f"-\\frac{{{clean_num}}}{{{rhs_den}}}"
                else:
                    rhs_str = f"\\frac{{{rhs_num}}}{{{rhs_den}}}" if rhs_num else f"\\frac{{1}}{{{rhs_den}}}"
                if rhs_str == "\\frac{}{}":
                    rhs_str = "\\frac{1}{" + rhs_den + "}"
            else:
                rhs_str = rhs_num
            if rhs_str:
                sign = "+"
                for item in non_primary:
                    cv = item["coeff_value"]
                    if cv is not None and cv < 0:
                        sign = "-"
                        break
                rhs_terms.append((sign, rhs_str))

    lhs = " + ".join(lhs_terms)

    rhs_parts = []
    for i, (sign, term_str) in enumerate(rhs_terms):
        if i == 0:
            rhs_parts.append(f"- {term_str}" if sign == "-" else term_str)
        else:
            rhs_parts.append(f"{sign} {term_str}")

    rhs = " ".join(rhs_parts)
    return f"{lhs} = {rhs}"


# ── Common queries ─────────────────────────────────────

def get_formula_detail(conn, formula_id):
    return conn.execute(
        "SELECT *, json_extract(name, '$.en-us') AS name_en,"
        " json_extract(branch, '$.en-us') AS branch_en,"
        " json_extract(topic, '$.en-us') AS topic_en,"
        " json_extract(description, '$.en-us') AS description_en"
        " FROM formulas WHERE id = ?",
        (formula_id,),
    ).fetchone()


def get_formula_items(conn, formula_id):
    return conn.execute("""
        SELECT fi.*, v.latex AS var_latex
        FROM formula_items fi
        LEFT JOIN variables v ON v.id = fi.variable_id
        WHERE fi.formula_id = ?
        ORDER BY fi.term, fi.is_primary DESC, fi.sort_order
    """, (formula_id,)).fetchall()


def get_formula_conditions(conn, formula_id):
    return conn.execute("""
        SELECT c.default_on, json_extract(c.name, '$.en-us') AS name_en,
               c.replacement_formula_id,
               json_extract(f2.name, '$.en-us') AS replacement_name
        FROM conditions c
        JOIN formulas f2 ON f2.id = c.replacement_formula_id
        WHERE c.formula_id = ?
        ORDER BY c.sort_order
    """, (formula_id,)).fetchall()


def get_formula_relations(conn, formula_id):
    return conn.execute("""
        SELECT fr.relation_type, fr.related_id,
               json_extract(f2.name, '$.en-us') AS related_name
        FROM formula_relations fr
        JOIN formulas f2 ON f2.id = fr.related_id
        WHERE fr.formula_id = ?
        ORDER BY fr.relation_type
    """, (formula_id,)).fetchall()


def get_formula_variables(conn, formula_id):
    return conn.execute("""
        SELECT DISTINCT v.id, v.name, v.latex, json_extract(v.name, '$.en-us') AS name_en,
               v.si_unit, v.dim_M, v.dim_L, v.dim_T, v.dim_I, v.dim_Θ, v.dim_N, v.dim_J
        FROM formula_items fi
        JOIN variables v ON v.id = fi.variable_id
        WHERE fi.formula_id = ?
        ORDER BY v.id
    """, (formula_id,)).fetchall()


def get_variable_detail(conn, variable_id):
    return conn.execute(
        "SELECT *, json_extract(name, '$.en-us') AS name_en,"
        " json_extract(description, '$.en-us') AS description_en"
        " FROM variables WHERE id = ?",
        (variable_id,),
    ).fetchone()


def get_variable_units(conn, variable_id):
    return conn.execute("""
        SELECT u.*, json_extract(u.name, '$.en-us') AS name_en
        FROM units u WHERE u.variable_id = ?
        ORDER BY u.si_unit DESC, u.unit_system
    """, (variable_id,)).fetchall()


def get_variable_formulas(conn, variable_id):
    return conn.execute("""
        SELECT DISTINCT f.id, json_extract(f.name, '$.en-us') AS name_en,
               json_extract(f.branch, '$.en-us') AS branch_en,
               json_extract(f.topic, '$.en-us') AS topic_en,
               f.difficulty
        FROM formula_items fi
        JOIN formulas f ON f.id = fi.formula_id
        WHERE fi.variable_id = ?
        ORDER BY f.branch, f.topic, f.difficulty, f.id
    """, (variable_id,)).fetchall()


def get_unit_detail(conn, unit_id):
    return conn.execute("""
        SELECT u.*, json_extract(v.name, '$.en-us') AS var_name,
               json_extract(u.name, '$.en-us') AS name_en
        FROM units u JOIN variables v ON v.id = u.variable_id
        WHERE u.id = ?
    """, (unit_id,)).fetchone()


def search(conn, query, limit=30):
    """Full-text search across formulas, variables, and units."""
    terms = " OR ".join(f'"{w}"*' for w in query.split())

    formulas = conn.execute("""
        SELECT 'formula' AS kind, fts.formula_id AS id,
               f.name AS name,
               json_extract(f.name, '$.en-us') AS name_en,
               json_extract(f.branch, '$.en-us') AS branch_en,
               json_extract(f.topic, '$.en-us') AS topic_en,
               f.difficulty, NULL AS extra
        FROM formula_fts fts
        JOIN formulas f ON f.id = fts.formula_id
        WHERE formula_fts MATCH ?
        ORDER BY rank
        LIMIT ?
    """, (terms, limit)).fetchall()

    variables = conn.execute("""
        SELECT 'variable' AS kind, fts.variable_id AS id,
               v.name AS name,
               fts.name AS name_en, NULL, NULL, NULL,
               v.latex AS extra
        FROM variable_fts fts
        JOIN variables v ON v.id = fts.variable_id
        WHERE variable_fts MATCH ?
        LIMIT ?
    """, (terms, limit)).fetchall()

    units = conn.execute("""
        SELECT 'unit' AS kind, fts.unit_id AS id,
               u.name AS name,
               fts.name AS name_en, NULL, NULL, NULL,
               fts.symbol AS extra
        FROM unit_fts fts
        JOIN units u ON u.id = fts.unit_id
        WHERE unit_fts MATCH ?
        LIMIT ?
    """, (terms, limit)).fetchall()

    return list(formulas) + list(variables) + list(units)


def get_base_units(conn):
    """Return SI base units with their variable info."""
    return conn.execute("""
         SELECT u.id, u.symbol, u.name, json_extract(u.name, '$.en-us') AS name_en,
               v.id AS variable_id, v.latex AS var_latex,
               v.name AS var_name_raw, json_extract(v.name, '$.en-us') AS var_name
        FROM units u
        JOIN variables v ON v.id = u.variable_id
        WHERE u.si_unit = 1 AND v.id IN ('length','mass','temperature','period',
              'electric_current','amount','luminous_intensity')
        ORDER BY v.id
    """).fetchall()


def get_all_variables(conn):
    """Return all variables with si_unit parsed and dimensions."""
    return conn.execute("""
        SELECT v.id, v.name, v.latex, json_extract(v.name, '$.en-us') AS name_en,
               v.si_unit,
               v.dim_M, v.dim_L, v.dim_T, v.dim_I, v.dim_Θ, v.dim_N, v.dim_J
        FROM variables v
        ORDER BY v.id
    """).fetchall()


def get_all_formulas(conn):
    """Return all formulas."""
    return conn.execute("""
        SELECT f.id, f.name, json_extract(f.name, '$.en-us') AS name_en,
               f.science, json_extract(f.science, '$.en-us') AS science_en,
               f.branch, json_extract(f.branch, '$.en-us') AS branch_en,
               f.topic, json_extract(f.topic, '$.en-us') AS topic_en,
               f.difficulty
        FROM formulas f
        ORDER BY f.science, f.branch, f.topic, f.difficulty, f.id
    """).fetchall()


def get_formulas_by_science(conn):
    """Return formulas grouped by science → branch → topic."""
    rows = conn.execute("""
        SELECT f.id, f.name, json_extract(f.name, '$.en-us') AS name_en,
               f.science, json_extract(f.science, '$.en-us') AS science_en,
               f.branch, json_extract(f.branch, '$.en-us') AS branch_en,
               f.topic, json_extract(f.topic, '$.en-us') AS topic_en,
               f.difficulty
        FROM formulas f
        ORDER BY f.science, f.branch, f.topic, f.difficulty, f.id
    """).fetchall()
    by_science = defaultdict(lambda: defaultdict(list))
    for r in rows:
        by_science[r["science_en"] or "Other"][r["branch_en"] or "Uncategorised"].append(r)
    return dict(by_science)



def render_si_unit_latex(si_unit_json):
    """Render si_unit JSON as siunitx LaTeX string."""
    parts = parse_si_unit_json(si_unit_json)
    if not parts:
        return ""
    num_parts = []
    den_parts = []
    for uid, exp in parts:
        # Map unit IDs to siunitx macros
        symbol_map = {
            "meter": "\\meter", "kilogram": "\\kilogram", "second": "\\second",
            "kelvin": "\\kelvin", "newton": "\\newton", "ohm": "\\ohm",
            "degree_celsius": "\\degreeCelsius",
        }
        s = symbol_map.get(uid, uid.replace("_", " "))
        if exp >= 0:
            num_parts.append((s, exp))
        else:
            den_parts.append((s, -exp))
    num_str = _siunitx_group(num_parts)
    if not den_parts:
        return f"\\si{{{num_str}}}"
    den_str = _siunitx_group(den_parts)
    return f"\\si{{{num_str}\\per{den_str}}}"


def _siunitx_group(parts):
    """Render a list of (symbol, exponent) as siunitx product."""
    items = []
    for sym, exp in parts:
        if exp == 1:
            items.append(sym)
        else:
            items.append(f"{sym}^{{{fmt_num(exp)}}}")
    return "\\,".join(items)

TABLE_ORDER = ["variables", "units", "formulas", "formula_items", "conditions", "formula_relations"]

TABLE_COLUMNS = {
    "formulas": [
        "id", "name", "science", "branch", "topic", "difficulty",
        "description", "links",
    ],
    "formula_items": [
        "formula_id", "term", "is_primary", "sort_order",
        "coeff_value", "coeff_special", "coeff_exponent",
        "variable_id", "var_exponent", "label",
        "latex_prefix", "latex_suffix", "latex_override",
    ],
    "conditions": [
        "name", "formula_id", "replacement_formula_id", "default_on", "sort_order",
    ],
    "formula_relations": [
        "formula_id", "related_id", "relation_type",
    ],
    "variables": [
        "id", "name", "latex", "science", "branch", "topic",
        "difficulty", "description", "links", "si_unit",
        "dim_M", "dim_L", "dim_T", "dim_I", "dim_Θ", "dim_N", "dim_J",
    ],
    "units": [
        "id", "variable_id", "symbol", "name", "factor_to_si", "offset",
        "si_unit", "unit_system",
    ],
}

TABLE_INSERT = {
    "formulas": """INSERT OR REPLACE INTO formulas
        (id, name, science, branch, topic, difficulty, description, links)
        VALUES (?,?,?,?,?,?,?,?)""",
    "formula_items": """INSERT OR REPLACE INTO formula_items
        (formula_id, term, is_primary, sort_order, coeff_value, coeff_special,
         coeff_exponent, variable_id, var_exponent, label,
         latex_prefix, latex_suffix, latex_override)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)""",
    "conditions": """INSERT OR REPLACE INTO conditions
        (name, formula_id, replacement_formula_id, default_on, sort_order)
        VALUES (?,?,?,?,?)""",
    "formula_relations": """INSERT OR REPLACE INTO formula_relations
        (formula_id, related_id, relation_type)
        VALUES (?,?,?)""",
    "variables": """INSERT OR REPLACE INTO variables
        (id, name, latex, science, branch, topic,
         difficulty, description, links, si_unit,
         dim_M, dim_L, dim_T, dim_I, dim_Θ, dim_N, dim_J)
        VALUES (?,?,?,?,?,?,?,?,?,?, ?,?,?,?,?,?,?)""",
    "units": """INSERT OR REPLACE INTO units
        (id, variable_id, symbol, name, factor_to_si, offset, si_unit, unit_system)
        VALUES (?,?,?,?,?,?,?,?)""",
}


def export_csv(conn):
    """Export all tables to a single CSV string with section headers."""
    buf = StringIO()
    for table in TABLE_ORDER:
        cols = TABLE_COLUMNS[table]
        buf.write(f"=== {table} ===\n")
        w = csv.writer(buf)
        w.writerow(cols)
        rows = conn.execute(f"SELECT {','.join(cols)} FROM {table} ORDER BY rowid").fetchall()
        for r in rows:
            w.writerow([r[c] for c in cols])
        buf.write("\n")
    return buf.getvalue()


def _table_rows(conn):
    """Yield (table_name, headers, rows) for all tables."""
    for table in TABLE_ORDER:
        cols = TABLE_COLUMNS[table]
        rows = conn.execute(f"SELECT {','.join(cols)} FROM {table} ORDER BY rowid").fetchall()
        yield table, cols, [[r[c] for c in cols] for r in rows]


def export_csv_dir(conn, directory):
    """Export each table to a separate CSV file in directory."""
    p = Path(directory)
    p.mkdir(parents=True, exist_ok=True)
    for table, cols, rows in _table_rows(conn):
        path = p / f"{table}.csv"
        with open(path, "w", newline="") as f:
            w = csv.writer(f)
            w.writerow(cols)
            w.writerows(rows)


def export_xlsx(conn, path):
    """Export all tables as sheets in an XLSX workbook."""
    from openpyxl import Workbook
    wb = Workbook()
    first = True
    for table, cols, rows in _table_rows(conn):
        if first:
            ws = wb.active
            ws.title = table[:31]  # sheet name max 31 chars
            first = False
        else:
            ws = wb.create_sheet(title=table[:31])
        ws.append(cols)
        for row in rows:
            ws.append(row)
    wb.save(path)


def export_ods(conn, path):
    """Export all tables as sheets in an ODS spreadsheet."""
    from odf.opendocument import OpenDocumentSpreadsheet
    from odf.table import Table, TableRow, TableCell
    from odf.text import P
    doc = OpenDocumentSpreadsheet()
    for table, cols, rows in _table_rows(conn):
        tbl = Table(name=table[:31])
        doc.spreadsheet.addElement(tbl)
        for row_data in [cols] + rows:
            tr = TableRow()
            for val in row_data:
                tc = TableCell()
                p = P(text=str(val) if val is not None else "")
                tc.addElement(p)
                tr.addElement(tc)
            tbl.addElement(tr)
    doc.save(path)


def _insert_rows(conn, table, headers, rows):
    """Insert rows from one table, updating counts dict."""
    expected = TABLE_COLUMNS[table]
    col_indices = [headers.index(col) for col in expected if col in headers]
    sql = TABLE_INSERT[table]
    cleaned = []
    for row in rows:
        if not row:
            continue
        vals = [row[i] if i < len(row) else "" for i in col_indices]
        cleaned.append(tuple(None if v == "" else v for v in vals))
    if cleaned:
        conn.executemany(sql, cleaned)
    return len(cleaned)


def import_csv_dir(conn, directory):
    """Import tables from per-table CSV files in a directory."""
    p = Path(directory)
    return _import_worksheets(conn, {
        table: _read_csv(p / f"{table}.csv")
        for table in TABLE_ORDER
        if (p / f"{table}.csv").exists()
    })


def _read_csv(path):
    """Read a CSV file, returning (headers, rows)."""
    with open(path, newline="") as f:
        reader = csv.reader(f)
        headers = next(reader, None)
        rows = list(reader)
    return headers, rows


def import_xlsx(conn, path):
    """Import tables from XLSX workbook sheets."""
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True)
    sheets = {}
    for name in wb.sheetnames:
        ws = wb[name]
        rows = list(ws.iter_rows(values_only=True))
        if rows:
            sheets[name] = (list(rows[0]), [list(r) for r in rows[1:]])
    wb.close()
    return _import_worksheets(conn, sheets)


def import_ods(conn, path):
    """Import tables from ODS spreadsheet sheets."""
    from odf.opendocument import load
    from odf.table import Table, TableRow
    from odf.text import P
    doc = load(path)
    sheets = {}
    for table_elem in doc.getElementsByType(Table):
        name = table_elem.getAttribute("name")
        rows = []
        for row_elem in table_elem.getElementsByType(TableRow):
            cells = []
            for cell in row_elem.childNodes:
                cell_text = None
                for p in cell.getElementsByType(P):
                    texts = []
                    for n in p.childNodes:
                        if hasattr(n, "data"):
                            texts.append(n.data)
                    cell_text = "".join(texts)
                    break
                cells.append(cell_text)
            if any(c is not None and c.strip() for c in cells):
                rows.append(cells)
        if rows:
            sheets[name] = (list(rows[0]), rows[1:])
    return _import_worksheets(conn, sheets)


def _import_worksheets(conn, sheets):
    """Import from a dict of {table_name: (headers, rows)}. Returns counts."""
    counts = {}
    # sheets may have truncated names — match by prefix
    table_map = {}
    for name in sheets:
        for t in TABLE_ORDER:
            if t.startswith(name) or name.startswith(t):
                table_map[t] = sheets[name]
                break
    for table in TABLE_ORDER:
        if table in table_map:
            headers, rows = table_map[table]
            n = _insert_rows(conn, table, headers, rows)
            counts[table] = n
    conn.commit()
    return counts


def import_csv(conn, csv_str):
    """Import tables from a CSV string with section headers. Returns counts."""
    counts = {}
    reader = csv.reader(StringIO(csv_str))
    current_table = None
    header_row = None
    col_indices = []
    rows_buffer = []

    def flush():
        if not current_table or not rows_buffer:
            return
        sql = TABLE_INSERT[current_table]
        expected = TABLE_COLUMNS[current_table]
        cleaned = []
        for row in rows_buffer:
            # map CSV columns to expected columns by index
            vals = [row[i] if i < len(row) else "" for i in col_indices]
            # convert empty strings to None for nullable columns
            cleaned.append(tuple(None if v == "" else v for v in vals))
        conn.executemany(sql, cleaned)
        counts[current_table] = len(rows_buffer)
        rows_buffer.clear()

    for row in reader:
        if not row or all(c.strip() == "" for c in row):
            continue
        if row[0].startswith("=== ") and row[0].endswith(" ==="):
            flush()
            current_table = row[0][4:-4].strip()
            header_row = None
            col_indices = []
            continue
        if current_table and header_row is None:
            header_row = row
            expected = TABLE_COLUMNS.get(current_table, [])
            col_indices = [header_row.index(col) for col in expected if col in header_row]
            continue
        if current_table and col_indices:
            rows_buffer.append(tuple(row))

    flush()
    conn.commit()
    return counts

